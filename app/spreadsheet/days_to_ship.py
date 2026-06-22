"""Set 備貨天數 (days to ship) to 1 for every product row.

PHASE 2 — scaffolded. The exact header name and the data-row layout must be
confirmed against a real export before this is trusted. openpyxl handles the
.xlsx Shopee normally exports; if the export is legacy .xls you'll need xlrd
(read) + xlwt (write) instead.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from app.shopee import selectors as S

log = logging.getLogger(__name__)

_CANDIDATE_HEADERS = [
    S.XLS_HEADER_DAYS_TO_SHIP,
    S.XLS_HEADER_DAYS_TO_SHIP_ALT,
]


def _find_header_col(ws) -> int | None:
    """Scan row 1 for a days-to-ship header. Returns 1-based column index."""
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        if cell.value in _CANDIDATE_HEADERS:
            return col_idx
    return None


def set_all_to_one(xlsx_path: Path) -> int:
    """Overwrite every data row's 備貨天數 with 1. Returns rows changed."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    col = _find_header_col(ws)
    if col is None:
        raise RuntimeError(
            f"找不到備貨天數欄位標題 in {xlsx_path}. "
            f"Expected one of {_CANDIDATE_HEADERS} — confirm the real header."
        )

    changed = 0
    for row in ws.iter_rows(min_row=2):  # skip header
        cell = row[col - 1]
        cell.value = S.TARGET_DAYS_TO_SHIP
        changed += 1

    wb.save(xlsx_path)
    log.info("set %d product rows to %d days-to-ship", changed, S.TARGET_DAYS_TO_SHIP)
    return changed
